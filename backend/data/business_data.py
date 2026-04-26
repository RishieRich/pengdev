"""
Load dashboard data from the root-level input folder.
The dashboard should only show KPIs when the required workbooks exist.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[2]
INPUT_DIR = PROJECT_ROOT / "input"
FY_START = date(2025, 4, 1)
FY_END = date(2026, 3, 31)
MONTHS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]


def _empty_data(message: str, missing_files: list[str] | None = None) -> dict[str, Any]:
    return {
        "available": False,
        "message": message,
        "missingFiles": missing_files or [],
        "inputDir": str(INPUT_DIR),
        "company": {
            "label": "Pawan Engineering",
            "workbookName": "",
            "address": "",
            "udyam": "",
            "period": "FY 2025-26 - Apr 2025 to Mar 2026",
        },
        "headline": None,
        "monthly": [],
        "topCustomers": [],
        "topVendors": [],
        "topProducts": [],
        "topMaterials": [],
        "alerts": [],
        "dataNotes": [],
    }


def _find_required_files() -> tuple[Path | None, Path | None, list[str]]:
    if not INPUT_DIR.exists():
        return None, None, ["input folder"]

    xlsx_files = [p for p in INPUT_DIR.glob("*.xlsx") if not p.name.startswith("~$")]
    sale_file = next((p for p in xlsx_files if "sale" in p.name.lower()), None)
    purchase_file = next((p for p in xlsx_files if "purchase" in p.name.lower()), None)

    missing = []
    if sale_file is None:
        missing.append("Sale 25-26.xlsx")
    if purchase_file is None:
        missing.append("Purchase 25-26.xlsx")
    return sale_file, purchase_file, missing


def _is_fy_date(value: Any) -> bool:
    if isinstance(value, datetime):
        value = value.date()
    return isinstance(value, date) and FY_START <= value <= FY_END


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _range_for_period(period: str | None) -> tuple[date, date]:
    ranges = {
        "fy": (FY_START, FY_END),
        "q1": (date(2025, 4, 1), date(2025, 6, 30)),
        "q2": (date(2025, 7, 1), date(2025, 9, 30)),
        "q3": (date(2025, 10, 1), date(2025, 12, 31)),
        "q4": (date(2026, 1, 1), FY_END),
        "h1": (date(2025, 4, 1), date(2025, 9, 30)),
        "h2": (date(2025, 10, 1), FY_END),
        "last90": (FY_END - timedelta(days=89), FY_END),
    }
    return ranges.get(period or "fy", ranges["fy"])


def _in_range(value: date, start: date, end: date) -> bool:
    return start <= value <= end


def _matches_query(values: list[str], query: str) -> bool:
    if not query:
        return True
    haystack = " ".join(values).lower()
    return query.lower() in haystack


def _clean_name(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def _pct(value: float, total: float) -> float:
    return round((value / total * 100), 1) if total else 0


def _top_rows(values: dict[str, float], counts: dict[str, int] | None, total: float, limit: int) -> list[dict[str, Any]]:
    rows = []
    for name, value in sorted(values.items(), key=lambda item: item[1], reverse=True)[:limit]:
        row = {"name": name, "value": round(value, 2), "share": _pct(value, total)}
        if counts is not None:
            row["lines"] = counts.get(name, 0)
        rows.append(row)
    return rows


def _read_company_meta(workbook: openpyxl.Workbook) -> dict[str, str]:
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(min_row=1, max_row=7, values_only=True))
    company = _clean_name(rows[0][0]) if rows else "Pawan Engineering"
    address_1 = _clean_name(rows[1][0]) if len(rows) > 1 else ""
    address_2 = _clean_name(rows[2][0]) if len(rows) > 2 else ""
    udyam = _clean_name(rows[3][0]).replace("UDYAM :", "").strip() if len(rows) > 3 else ""
    return {
        "label": "Pawan Engineering",
        "workbookName": company,
        "address": _clean_name(f"{address_1} {address_2}"),
        "udyam": udyam,
        "period": "FY 2025-26 - Apr 2025 to Mar 2026",
    }


def _read_sales_register(workbook: openpyxl.Workbook) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    sheet = workbook["Sales Register"]
    invoices = []
    lines = []
    current_date = None
    current_customer = ""

    for row in sheet.iter_rows(min_row=9, values_only=True):
        row_date = row[0]
        value = row[7] if len(row) > 7 else None

        if _is_fy_date(row_date) and isinstance(value, (int, float)):
            current_date = row_date.date()
            current_customer = _clean_name(row[1])
            invoices.append({"date": current_date, "party": current_customer, "value": float(value)})
            continue
        if isinstance(row_date, datetime):
            current_date = None
            current_customer = ""
            continue

        product = _clean_name(row[1] if len(row) > 1 else "")
        if current_date and product and product.lower() != "grand total" and isinstance(value, (int, float)):
            lines.append({
                "date": current_date,
                "party": current_customer,
                "product": product,
                "value": float(value),
            })

    return invoices, lines


def _read_purchase_register(workbook: openpyxl.Workbook) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    sheet = workbook["Purchase Register"]
    invoices = []
    lines = []
    current_date = None
    current_vendor = ""

    for row in sheet.iter_rows(min_row=9, values_only=True):
        row_date = row[0]
        value = row[4] if len(row) > 4 else None

        if _is_fy_date(row_date) and isinstance(value, (int, float)):
            current_date = row_date.date()
            current_vendor = _clean_name(row[1])
            invoices.append({"date": current_date, "party": current_vendor, "value": float(value)})
            continue
        if isinstance(row_date, datetime):
            current_date = None
            current_vendor = ""
            continue

        material = _clean_name(row[1] if len(row) > 1 else "")
        if current_date and material and material.lower() != "grand total" and isinstance(value, (int, float)):
            lines.append({
                "date": current_date,
                "party": current_vendor,
                "material": material,
                "value": float(value),
            })

    return invoices, lines


def _group_by_party(records: list[dict[str, Any]]) -> tuple[dict[str, float], dict[str, int]]:
    values: dict[str, float] = defaultdict(float)
    counts: dict[str, int] = defaultdict(int)
    for record in records:
        values[record["party"]] += record["value"]
        counts[record["party"]] += 1
    return values, counts


def _group_by_key(records: list[dict[str, Any]], key: str) -> dict[str, float]:
    values: dict[str, float] = defaultdict(float)
    for record in records:
        name = record.get(key, "")
        if name:
            values[name] += record["value"]
    return values


def _filter_records(records: list[dict[str, Any]], start: date, end: date, query: str, keys: list[str]) -> list[dict[str, Any]]:
    filtered = []
    for record in records:
        if not _in_range(record["date"], start, end):
            continue
        if not _matches_query([str(record.get(key, "")) for key in keys], query):
            continue
        filtered.append(record)
    return filtered


def _read_register(workbook: openpyxl.Workbook, sheet_name: str, value_col: int) -> tuple[list[dict[str, Any]], dict[str, float], dict[str, int]]:
    sheet = workbook[sheet_name]
    records = []
    by_party: dict[str, float] = defaultdict(float)
    counts: dict[str, int] = defaultdict(int)

    for row in sheet.iter_rows(min_row=9, values_only=True):
        row_date = row[0]
        value = row[value_col] if len(row) > value_col else None
        if not _is_fy_date(row_date) or not isinstance(value, (int, float)):
            continue

        party = _clean_name(row[1])
        records.append({"date": row_date.date(), "party": party, "value": float(value)})
        by_party[party] += float(value)
        counts[party] += 1

    return records, by_party, counts


def _build_alerts(headline: dict[str, Any], top_customers: list[dict[str, Any]], top_vendors: list[dict[str, Any]], monthly: list[dict[str, Any]]) -> list[dict[str, str]]:
    alerts = []
    if top_customers and top_customers[0]["share"] >= 30:
        alerts.append({
            "sev": "high",
            "title": "Customer concentration zyada hai",
            "detail": f"{top_customers[0]['name']} akele {top_customers[0]['share']}% sales dete hain. Top 5 customers milke {headline['top5CustShare']}%.",
        })
    if headline["top3VendShare"] >= 60:
        alerts.append({
            "sev": "high",
            "title": "Vendor dependency check karein",
            "detail": f"Top 3 vendors milke {headline['top3VendShare']}% purchases karte hain. Backup suppliers validate karein.",
        })
    march = next((m for m in monthly if m["m"] == "Mar"), None)
    if march and march["gp"] < 0:
        alerts.append({
            "sev": "high",
            "title": "March 2026 mein gross loss",
            "detail": "March gross profit negative hai. Procurement timing ya billing cutoff accounts se confirm karein.",
        })
    return alerts


def load_business_data(
    period: str | None = "fy",
    start_date: str | None = None,
    end_date: str | None = None,
    q: str | None = None,
) -> dict[str, Any]:
    sale_file, purchase_file, missing = _find_required_files()
    if missing:
        return _empty_data(
            "Please check the input folder. Required data files are missing, so there are no KPIs to show.",
            missing,
        )

    try:
        purchase_workbook = openpyxl.load_workbook(purchase_file, read_only=True, data_only=True)
        try:
            company = _read_company_meta(purchase_workbook)
            sales_records, sales_lines = _read_sales_register(purchase_workbook)
            purchase_records, purchase_lines = _read_purchase_register(purchase_workbook)
        finally:
            purchase_workbook.close()
    except Exception as exc:
        return _empty_data(f"Input files could not be read: {exc}", [])

    range_start, range_end = _range_for_period(period)
    range_start = _parse_date(start_date) or range_start
    range_end = _parse_date(end_date) or range_end
    if range_start > range_end:
        range_start, range_end = range_end, range_start
    query = (q or "").strip()

    sales_records = _filter_records(sales_records, range_start, range_end, query, ["party"])
    sales_lines = _filter_records(sales_lines, range_start, range_end, query, ["party", "product"])
    purchase_records = _filter_records(purchase_records, range_start, range_end, query, ["party"])
    purchase_lines = _filter_records(purchase_lines, range_start, range_end, query, ["party", "material"])

    sales_total = round(sum(r["value"] for r in sales_records), 2)
    purchases_total = round(sum(r["value"] for r in purchase_records), 2)
    if sales_total <= 0 and purchases_total <= 0:
        return {
            "available": True,
            "message": "No matching records found for this filter.",
            "missingFiles": [],
            "inputDir": str(INPUT_DIR),
            "sourceFiles": [sale_file.name, purchase_file.name],
            "filters": {"period": period or "fy", "startDate": range_start.isoformat(), "endDate": range_end.isoformat(), "q": query},
            "company": company,
            "headline": {
                "sales": 0,
                "purchases": 0,
                "grossProfit": 0,
                "grossMarginPct": 0,
                "salesCount": 0,
                "purchaseCount": 0,
                "customers": 0,
                "vendors": 0,
                "products": 0,
                "top5CustShare": 0,
                "top3VendShare": 0,
            },
            "monthly": [{"m": month, "sales": 0, "purchases": 0, "gp": 0} for month in MONTHS],
            "topCustomers": [],
            "topVendors": [],
            "topProducts": [],
            "topMaterials": [],
            "alerts": [],
            "dataNotes": ["No rows matched the selected tenure or search filter."],
        }

    sales_by_customer, sales_counts = _group_by_party(sales_records)
    purchase_by_vendor, purchase_counts = _group_by_party(purchase_records)
    sales_by_product = _group_by_key(sales_lines, "product")
    purchase_by_material = _group_by_key(purchase_lines, "material")

    customers = len(sales_by_customer)
    vendors = len(purchase_by_vendor)

    monthly_sales: dict[str, float] = defaultdict(float)
    monthly_purchases: dict[str, float] = defaultdict(float)
    for row in sales_records:
        monthly_sales[row["date"].strftime("%b")] += row["value"]
    for row in purchase_records:
        monthly_purchases[row["date"].strftime("%b")] += row["value"]

    monthly = []
    for month in MONTHS:
        sales = round(monthly_sales[month], 2)
        purchases = round(monthly_purchases[month], 2)
        monthly.append({"m": month, "sales": sales, "purchases": purchases, "gp": round(sales - purchases, 2)})

    gross_profit = round(sales_total - purchases_total, 2)
    top_customers = _top_rows(sales_by_customer, sales_counts, sales_total, 6)
    top_vendors = _top_rows(purchase_by_vendor, purchase_counts, purchases_total, 5)
    top_products = _top_rows(sales_by_product, None, sales_total, 5)
    top_materials = _top_rows(purchase_by_material, None, purchases_total, 5)

    headline = {
        "sales": sales_total,
        "purchases": purchases_total,
        "grossProfit": gross_profit,
        "grossMarginPct": round((gross_profit / sales_total * 100), 2) if sales_total else 0,
        "salesCount": len(sales_records),
        "purchaseCount": len(purchase_records),
        "customers": customers,
        "vendors": vendors,
        "products": len(sales_by_product),
        "top5CustShare": round(sum(c["share"] for c in top_customers[:5]), 1),
        "top3VendShare": round(sum(v["share"] for v in top_vendors[:3]), 1),
    }

    return {
        "available": True,
        "message": "",
        "missingFiles": [],
        "inputDir": str(INPUT_DIR),
        "sourceFiles": [sale_file.name, purchase_file.name],
        "filters": {"period": period or "fy", "startDate": range_start.isoformat(), "endDate": range_end.isoformat(), "q": query},
        "company": company,
        "headline": headline,
        "monthly": monthly,
        "topCustomers": top_customers,
        "topVendors": top_vendors,
        "topProducts": top_products,
        "topMaterials": top_materials,
        "alerts": _build_alerts(headline, top_customers, top_vendors, monthly),
        "dataNotes": [
            "KPIs are loaded from the root input folder at request time.",
            "Headline sales and purchases use dated FY rows from the registers and exclude grand totals.",
            "Purchase rows after 31-Mar-2026 are excluded from FY 2025-26 KPIs.",
            "Gross profit is trading-level only and excludes overheads, tax, finance cost, and payroll.",
        ],
    }


PE_DATA = load_business_data()
